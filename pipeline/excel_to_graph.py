#!/usr/bin/env python3
"""Convert a hierarchical TOM export (ARIS -> Excel/CSV) into graph nodes & edges.

This is the production-hardened successor to the POC's conversion script. It:
  - reads a flat hierarchy (code, name, level, parent_code, roles, controls),
  - attaches process-flow JSON to L1 nodes from a sidecar file,
  - validates the hierarchy (missing parents, level jumps, duplicates, cycles),
  - emits a graph JSON artifact (nodes + edges), and
  - prints a reconciliation report.

Optionally (`--load`) it upserts straight into the configured graph backend
(Neo4j dev / Cosmos Gremlin prod) via the backend's GraphStore abstraction.

Usage:
  python excel_to_graph.py --input sample_data/finance_tom_sample.csv \
      --flows sample_data/finance_flows.json \
      --output sample_data/finance_graph.generated.json
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Optional


def _read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:  # pragma: no cover
            sys.exit("openpyxl is required to read .xlsx files (pip install openpyxl)")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        return [
            {header[i]: ("" if c is None else str(c)) for i, c in enumerate(r)}
            for r in rows[1:]
            if any(c is not None for c in r)
        ]
    with path.open(newline="", encoding="utf-8-sig") as f:
        return [dict(r) for r in csv.DictReader(f)]


def _split(value: str) -> list[str]:
    return [v.strip() for v in (value or "").split(";") if v.strip()]


def _stable_code(prefix: str, name: str, metadata: dict[str, str]) -> str:
    basis = "|".join(
        [
            metadata.get("sector", ""),
            metadata.get("function", ""),
            metadata.get("technology", ""),
            name.strip().lower(),
        ]
    )
    digest = hashlib.sha1(basis.encode("utf-8")).hexdigest()[:10].upper()
    initials = "".join(w[0] for w in name.upper().split())[:6] or "ITEM"
    return f"{prefix}-{initials}-{digest}"


def _normalise_metadata(
    rows: list[dict[str, str]],
    sector: str = "",
    function: str = "",
    technology: str = "",
) -> dict[str, str]:
    root_name = next(
        ((r.get("name") or "").strip() for r in rows if str(r.get("level", "")).strip() == "0"),
        "",
    )
    return {
        "sector": sector.strip() or "Cross-sector",
        "function": function.strip() or root_name or "Unspecified",
        "technology": technology.strip() or "Tech-agnostic",
    }


def build_graph(
    rows: list[dict[str, str]],
    flows: dict[str, Any],
    metadata: Optional[dict[str, str]] = None,
) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    metadata = metadata or _normalise_metadata(rows)
    nodes: dict[str, dict[str, Any]] = {}
    edges: list[dict[str, str]] = []
    role_nodes: dict[str, dict[str, Any]] = {}
    control_nodes: dict[str, dict[str, Any]] = {}

    seen_codes: set[str] = set()
    for row in rows:
        code = (row.get("code") or "").strip()
        name = (row.get("name") or "").strip()
        if not code or not name:
            warnings.append(f"Skipped row with missing code/name: {row}")
            continue
        if code in seen_codes:
            warnings.append(f"Duplicate code '{code}' — keeping first occurrence")
            continue
        seen_codes.add(code)
        try:
            level = int(row.get("level", ""))
        except ValueError:
            warnings.append(f"Row '{code}' has non-integer level; defaulting to 0")
            level = 0

        node: dict[str, Any] = {
            "code": code,
            "name": name,
            "level": level,
            "type": "Process",
            **metadata,
        }
        if level == 1 and code in flows:
            node["process_flow_json"] = flows[code]
        nodes[code] = node

        # roles / controls
        for role in _split(row.get("roles", "")):
            rc = _stable_code("ROLE", role, metadata)
            role_nodes.setdefault(rc, {"code": rc, "name": role, "type": "Role", **metadata})
            edges.append({"from": code, "to": rc, "type": "PERFORMED_BY"})
        for ctrl in _split(row.get("controls", "")):
            cc = _stable_code("CTRL", ctrl, metadata)
            control_nodes.setdefault(cc, {"code": cc, "name": ctrl, "type": "Control", **metadata})
            edges.append({"from": code, "to": cc, "type": "HAS_CONTROL"})

    # hierarchy edges + validation
    for row in rows:
        code = (row.get("code") or "").strip()
        parent = (row.get("parent_code") or "").strip()
        if not code or code not in nodes:
            continue
        if not parent:
            if nodes[code]["level"] != 0:
                warnings.append(f"Node '{code}' (L{nodes[code]['level']}) has no parent")
            continue
        if parent not in nodes:
            warnings.append(f"Node '{code}' references missing parent '{parent}'")
            continue
        if nodes[parent]["level"] != nodes[code]["level"] - 1:
            warnings.append(
                f"Level jump: '{code}' (L{nodes[code]['level']}) under "
                f"'{parent}' (L{nodes[parent]['level']})"
            )
        edges.append({"from": parent, "to": code, "type": "HAS_SUB_PROCESS"})

    _detect_cycles(nodes, edges, warnings)

    flows_used = [c for c in flows if c in nodes]
    for c in flows:
        if c not in nodes:
            warnings.append(f"Flow defined for unknown code '{c}'")

    graph = {
        "metadata": {
            **metadata,
            "generated_by": "excel_to_graph.py",
            "flows_attached": flows_used,
        },
        "nodes": list(nodes.values()) + list(role_nodes.values()) + list(control_nodes.values()),
        "edges": edges,
    }
    return graph, warnings


def _detect_cycles(nodes, edges, warnings) -> None:
    children: dict[str, list[str]] = {}
    for e in edges:
        if e["type"] == "HAS_SUB_PROCESS":
            children.setdefault(e["from"], []).append(e["to"])
    WHITE, GREY, BLACK = 0, 1, 2
    color = {c: WHITE for c in nodes}

    def visit(n: str) -> bool:
        color[n] = GREY
        for m in children.get(n, []):
            if color.get(m) == GREY:
                return True
            if color.get(m) == WHITE and visit(m):
                return True
        color[n] = BLACK
        return False

    for c in list(nodes):
        if color[c] == WHITE and visit(c):
            warnings.append(f"Cycle detected in hierarchy near '{c}'")
            break


def reconciliation(graph: dict[str, Any]) -> str:
    levels = Counter(n["level"] for n in graph["nodes"] if n["type"] == "Process")
    types = Counter(n["type"] for n in graph["nodes"])
    edge_types = Counter(e["type"] for e in graph["edges"])
    lines = ["Reconciliation report", "----------------------"]
    lines.append("Process nodes by level: " + ", ".join(f"L{k}={levels[k]}" for k in sorted(levels)))
    lines.append("Nodes by type: " + ", ".join(f"{k}={v}" for k, v in types.items()))
    lines.append("Edges by type: " + ", ".join(f"{k}={v}" for k, v in edge_types.items()))
    return "\n".join(lines)


def main() -> None:
    ap = argparse.ArgumentParser(description="ARIS/Excel hierarchy -> graph nodes & edges")
    ap.add_argument("--input", required=True, type=Path)
    ap.add_argument("--flows", type=Path, help="Optional sidecar JSON of process flows by L1 code")
    ap.add_argument("--output", type=Path, default=Path("finance_graph.generated.json"))
    ap.add_argument("--sector", default="", help="Sector metadata applied to every node")
    ap.add_argument("--function", default="", help="Function/vertical metadata applied to every node")
    ap.add_argument("--technology", default="", help="Technology metadata applied to every node")
    ap.add_argument("--load", action="store_true", help="Upsert into the configured graph backend")
    args = ap.parse_args()

    rows = _read_rows(args.input)
    flows = json.loads(args.flows.read_text()) if args.flows and args.flows.exists() else {}
    metadata = _normalise_metadata(rows, args.sector, args.function, args.technology)
    graph, warnings = build_graph(rows, flows, metadata)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(graph, indent=2))
    print(f"Wrote {args.output} "
          f"({len(graph['nodes'])} nodes, {len(graph['edges'])} edges)")
    print(reconciliation(graph))
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")

    if args.load:
        sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "backend"))
        from app.clients.graph_store import get_graph_store

        store = get_graph_store()
        process_nodes = [
            {**n, "process_flow_json": json.dumps(n["process_flow_json"])}
            if "process_flow_json" in n else n
            for n in graph["nodes"]
        ]
        n_nodes = store.upsert_nodes(process_nodes)
        n_edges = store.upsert_edges(graph["edges"])
        store.close()
        print(f"\nLoaded into graph backend: {n_nodes} nodes, {n_edges} edges")


if __name__ == "__main__":
    main()
